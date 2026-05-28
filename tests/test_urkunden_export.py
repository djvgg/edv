# SPDX-FileCopyrightText: 2026 TOP Team Combat Control
# SPDX-License-Identifier: GPL-3.0-or-later

"""Tests für den Urkunden-Export (Datenquelle für den Serienbrief)."""

from openpyxl import load_workbook

from backend.services.urkunden_export_service import (
    COLUMNS,
    UrkundenExportService,
    compose_class_label,
    load_exported_keys,
    mark_exported,
)


class FakeDB:
    """Duck-typed Ersatz für DatabaseService — keine echte DB nötig."""

    def __init__(self, placements):
        # placements: {bracket_key: [placement-dicts]}
        self._placements = placements

    def get_bracket_placements(self, key):
        return self._placements.get(key, [])

    def get_completed_bracket_keys(self):
        return set(self._placements.keys())


# ----- Klassenlabel-Komposition ---------------------------------------

class TestComposeClassLabel:
    def test_canonical(self):
        assert compose_class_label('w | U13 | -40kg') == 'U13 -40kg weiblich'

    def test_male(self):
        assert compose_class_label('m | U15 | -50kg') == 'U15 -50kg männlich'

    def test_no_class_weight_omitted(self):
        assert compose_class_label('m | U18 | no-class') == 'U18 männlich'

    def test_empty_weight_omitted(self):
        assert compose_class_label('w | U11 | ') == 'U11 weiblich'

    def test_non_canonical_falls_back_to_raw(self):
        assert compose_class_label('QUARANTINE') == 'QUARANTINE'

    def test_unknown_gender_token_kept_verbatim(self):
        assert compose_class_label('x | U9 | -25kg') == 'U9 -25kg x'


# ----- Zeilenaufbau ----------------------------------------------------

def _placement(platz, vorname, nachname, verein=''):
    return {'platz': platz, 'vorname': vorname, 'nachname': nachname,
            'bracket_type': 'ko', 'verein': verein}


class TestBuildRows:
    def test_ordinal_and_klasse(self):
        db = FakeDB({'w | U13 | -40kg': [_placement(1, 'Lena', 'Müller', 'JC Worms')]})
        svc = UrkundenExportService(db=db)
        rows = svc.build_rows(['w | U13 | -40kg'])
        assert rows == [{
            'vorname': 'Lena', 'nachname': 'Müller', 'platz': '1.',
            'klasse': 'U13 -40kg weiblich', 'verein': 'JC Worms',
        }]

    def test_double_pool_yields_four_rows(self):
        # Doppelpool: 1., 2. und ZWEI dritte Plätze.
        db = FakeDB({'m | U13 | -34kg': [
            _placement(1, 'A', 'Eins'),
            _placement(2, 'B', 'Zwei'),
            _placement(3, 'C', 'Drei'),
            _placement(3, 'D', 'Vier'),
        ]})
        svc = UrkundenExportService(db=db)
        rows = svc.build_rows(['m | U13 | -34kg'])
        assert len(rows) == 4
        assert [r['platz'] for r in rows] == ['1.', '2.', '3.', '3.']

    def test_partial_selection_only_chosen_brackets(self):
        db = FakeDB({
            'm | U13 | -40kg': [_placement(1, 'A', 'Eins')],
            'w | U15 | -44kg': [_placement(1, 'B', 'Zwei')],
        })
        svc = UrkundenExportService(db=db)
        rows = svc.build_rows(['w | U15 | -44kg'])
        assert len(rows) == 1
        assert rows[0]['nachname'] == 'Zwei'

    def test_sorted_by_klasse_then_platz(self):
        db = FakeDB({
            'w | U15 | -44kg': [_placement(2, 'B', 'Zwei'), _placement(1, 'A', 'Eins')],
            'm | U13 | -40kg': [_placement(1, 'C', 'Drei')],
        })
        svc = UrkundenExportService(db=db)
        rows = svc.build_rows(['w | U15 | -44kg', 'm | U13 | -40kg'])
        klassen = [r['klasse'] for r in rows]
        # U13 sortiert vor U15; innerhalb U15: 1. vor 2.
        assert klassen == ['U13 -40kg männlich',
                           'U15 -44kg weiblich', 'U15 -44kg weiblich']
        assert [r['platz'] for r in rows[1:]] == ['1.', '2.']


# ----- XLSX-Ausgabe ----------------------------------------------------

class TestExport:
    def test_writes_header_and_rows(self, tmp_path):
        db = FakeDB({'m | U13 | -34kg': [
            _placement(1, 'A', 'Eins'), _placement(2, 'B', 'Zwei'),
            _placement(3, 'C', 'Drei'), _placement(3, 'D', 'Vier'),
        ]})
        out = tmp_path / 'urkunden.xlsx'
        svc = UrkundenExportService(db=db)
        result = svc.export(['m | U13 | -34kg'], out_path=str(out))

        assert result == {'rows': 4, 'brackets': 1, 'path': str(out)}
        wb = load_workbook(out)
        ws = wb.active
        assert ws.title == 'urkunden'
        assert [c.value for c in ws[1]] == COLUMNS
        assert ws.max_row == 5  # header + 4

    def test_rerun_overwrites(self, tmp_path):
        out = tmp_path / 'urkunden.xlsx'
        db1 = FakeDB({'m | U13 | -40kg': [_placement(1, 'A', 'Eins')]})
        UrkundenExportService(db=db1).export(['m | U13 | -40kg'], out_path=str(out))
        db2 = FakeDB({'w | U15 | -44kg': [
            _placement(1, 'B', 'Zwei'), _placement(2, 'C', 'Drei')]})
        UrkundenExportService(db=db2).export(['w | U15 | -44kg'], out_path=str(out))

        ws = load_workbook(out).active
        assert ws.max_row == 3  # header + 2 (alte Klasse weg)
        assert ws.cell(row=2, column=2).value == 'Zwei'


# ----- Sidecar-Marker --------------------------------------------------

class TestSidecar:
    def test_missing_file_is_empty(self, tmp_path):
        assert load_exported_keys(str(tmp_path / 'nope.json')) == set()

    def test_mark_and_load_roundtrip(self, tmp_path):
        path = str(tmp_path / 'printed.json')
        mark_exported(['m | U13 | -40kg'], path=path)
        mark_exported(['w | U15 | -44kg'], path=path)
        assert load_exported_keys(path) == {'m | U13 | -40kg', 'w | U15 | -44kg'}

    def test_mark_is_idempotent_union(self, tmp_path):
        path = str(tmp_path / 'printed.json')
        mark_exported(['m | U13 | -40kg'], path=path)
        result = mark_exported(['m | U13 | -40kg'], path=path)
        assert result == {'m | U13 | -40kg'}

    def test_corrupt_file_is_empty(self, tmp_path):
        path = tmp_path / 'printed.json'
        path.write_text('{not valid json', encoding='utf-8')
        assert load_exported_keys(str(path)) == set()
