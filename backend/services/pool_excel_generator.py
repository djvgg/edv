"""
Excel pool generator - adapts bracket_pdf_generator_v2 pool logic to Excel format.
"""

import logging
import os
from typing import List, Dict, Any, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utils.logging import get_logger

# Import existing pool logic
import sys
sys.path.insert(0, r'c:\UNI\TOP\edv_backend')
from frontend.utils.pool_renderer import _generate_fight_schedule
from frontend.styles import COLORS

logger = get_logger('pool_excel_generator')

# Import bracket drawing for finale brackets
try:
    from backend.services.bracket_excel_generator import BracketExcelGenerator
    BRACKET_GENERATOR_AVAILABLE = True
except ImportError:
    BRACKET_GENERATOR_AVAILABLE = False

# Excel styling - using frontend COLORS
HEADER_FILL = PatternFill(start_color=COLORS['accent_blue'].lstrip('#'), end_color=COLORS['accent_blue'].lstrip('#'), fill_type="solid")
HEADER_FONT = Font(bold=True, color=COLORS['text_primary'].lstrip('#'), size=10)
POOL_TITLE_FONT = Font(bold=True, size=12, color=COLORS['text_primary'].lstrip('#'))
SECTION_LABEL_FONT = Font(size=8, italic=True, color=COLORS['text_secondary'].lstrip('#'))
FIGHTER_FONT = Font(size=9, color=COLORS['text_primary'].lstrip('#'))
FIGHT_HEADER_FONT = Font(bold=True, size=8, color=COLORS['text_primary'].lstrip('#'))
FIGHT_HEADER_FILL = PatternFill(start_color=COLORS['bg_darker'].lstrip('#'), end_color=COLORS['bg_darker'].lstrip('#'), fill_type="solid")
NON_FIGHT_FILL = PatternFill(start_color=COLORS['border_light'].lstrip('#'), end_color=COLORS['border_light'].lstrip('#'), fill_type="solid")
RESULT_HEADER_FILL = PatternFill(start_color=COLORS['accent_blue'].lstrip('#'), end_color=COLORS['accent_blue'].lstrip('#'), fill_type="solid")
RESULT_HEADER_FONT = Font(bold=True, size=9, color=COLORS['text_primary'].lstrip('#'))
BORDER_THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


class PoolExcelGenerator:
    """Generate tournament pools in Excel format."""
    
    def __init__(self):
        self.logger = logger
    
    def generate_pools_excel(self,
                             output_path: str,
                             pools_data: List[Dict[str, Any]],
                             title: str = "Tournament Pools",
                             include_finale: bool = True) -> bool:
        """
        Generate Excel file with pool standings and optional finale bracket.
        
        Args:
            output_path: Path to save Excel file
            pools_data: List of pool dicts with 'pool_name', 'fighters'
            title: Workbook title
            include_finale: If True and 2+ pools, draw finale bracket with pool winners
            
        Returns:
            True if successful
        """
        try:
            os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Pools"
            
            # Set up sheet properties
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.orientation = 'landscape'
            
            # Main title
            ws['A1'] = title
            ws['A1'].font = Font(bold=True, size=14)
            
            # Column setup
            for col in range(1, 50):
                ws.column_dimensions[get_column_letter(col)].width = 12
            
            # Draw all pools and collect finalists
            current_row = 3
            pool_finalists = []  # Track (pool_name, 1st_place, 2nd_place) for each pool
            
            for pool_idx, pool_data in enumerate(pools_data):
                current_row = self._draw_pool_table(ws, pool_data, current_row)
                
                # For now, mark placeholder finalists (user will fill in rankings)
                pool_name = pool_data.get('pool_name', f'Pool {pool_idx + 1}')
                pool_finalists.append({
                    'pool_name': pool_name,
                    'rank_1': 'TBD',  # To be filled after pool scores entered
                    'rank_2': 'TBD'
                })
                
                current_row += 2  # Gap between pools
            
            # Draw finale bracket if 2+ pools
            if include_finale and len(pools_data) >= 2 and BRACKET_GENERATOR_AVAILABLE:
                finale_start_row = current_row + 1
                self._draw_finale_bracket(ws, pool_finalists, finale_start_row)
            
            # Save
            wb.save(output_path)
            self.logger.info(f"✓ Pools Excel file created: {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error creating pools Excel: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _draw_pool_table(self, ws, pool_data: Dict[str, Any], start_row: int) -> int:
        """
        Draw a pool round-robin table with fight scores.
        
        Layout:
            Pool Title
            Nr | Fighter | Club | Fight 1 | Fight 2 | ... | Points | Diff | Place
            ---|---------|------|---------|---------|-----|--------|------|-------
            
        Returns:
            Next available row after pool
        """
        fighters = pool_data.get('fighters', [])
        if not fighters:
            return start_row
        
        pool_name = pool_data.get('pool_name', 'Pool')
        fights_data = pool_data.get('fights', {})  # Dict mapping (idx1, idx2) -> {scores}
        num_fighters = len(fighters)
        
        # Pool title
        title_row = start_row
        ws[f'A{title_row}'] = pool_name
        ws[f'A{title_row}'].font = POOL_TITLE_FONT
        
        # Generate fight schedule
        fight_schedule = _generate_fight_schedule(num_fighters)
        num_fights = len(fight_schedule)
        
        # Build header row
        current_row = start_row + 2
        header_row = current_row
        
        col = 1
        # Fixed columns
        header_cells = [
            ('Nr', 'A'),
            ('Fighter', 'B'),
            ('Club', 'C'),
        ]
        
        for header_text, col_letter in header_cells:
            cell = ws[f'{col_letter}{header_row}']
            cell.value = header_text
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = BORDER_THIN
            ws.row_dimensions[header_row].height = 18
        
        col = 4  # Start fight columns at D
        # Fight columns
        for fight_idx in range(num_fights):
            col_letter = get_column_letter(col)
            # Show fight number as "F1", "F2", etc
            cell = ws[f'{col_letter}{header_row}']
            opponent_idx_1 = fight_schedule[fight_idx][0][0]
            opponent_idx_2 = fight_schedule[fight_idx][0][1]
            cell.value = f"F{fight_idx + 1}"
            cell.font = FIGHT_HEADER_FONT
            cell.fill = FIGHT_HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = BORDER_THIN
            col += 1
        
        # Result columns - Points, Diff, Place
        for header_text in ['Points', 'Diff', 'Place']:
            col_letter = get_column_letter(col)
            cell = ws[f'{col_letter}{header_row}']
            cell.value = header_text
            cell.font = RESULT_HEADER_FONT
            cell.fill = RESULT_HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = BORDER_THIN
            col += 1
        
        # Fighter rows - empty form for manual entry
        current_row += 1
        for fighter_idx, fighter in enumerate(fighters):
            row = current_row + fighter_idx
            
            # Parse fighter data (name, club) or dict
            if isinstance(fighter, (list, tuple)):
                name = fighter[0] if len(fighter) > 0 else f'Fighter {fighter_idx + 1}'
                club = fighter[1] if len(fighter) > 1 else ''
            elif isinstance(fighter, dict):
                name = fighter.get('Name', fighter.get('name', f'Fighter {fighter_idx + 1}'))
                club = fighter.get('Verein', fighter.get('club', ''))
            else:
                name = str(fighter)
                club = ''
            
            # Nr
            cell = ws[f'A{row}']
            cell.value = fighter_idx + 1
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = BORDER_THIN
            cell.font = FIGHTER_FONT
            ws.row_dimensions[row].height = 16
            
            # Name
            cell = ws[f'B{row}']
            cell.value = name
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = BORDER_THIN
            cell.font = FIGHTER_FONT
            
            # Club
            cell = ws[f'C{row}']
            cell.value = club
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = BORDER_THIN
            cell.font = Font(size=8, italic=True)
            
            # Fight results - empty for manual entry
            col = 4
            for fight_idx, fight_pair in enumerate(fight_schedule):
                opponent_idx_1 = fight_pair[0][0]
                opponent_idx_2 = fight_pair[0][1]
                
                # Check if this fighter participates in this fight
                is_in_fight = fighter_idx == opponent_idx_1 or fighter_idx == opponent_idx_2
                
                col_letter = get_column_letter(col)
                cell = ws[f'{col_letter}{row}']
                
                if is_in_fight:
                    # Empty cell for score entry
                    cell.value = ''
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                else:
                    # Non-fight cell - color it gray with "x"
                    cell.value = "x"
                    cell.fill = NON_FIGHT_FILL
                    cell.font = Font(size=7, color="999999")
                
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = BORDER_THIN
                col += 1
            
            # Results columns: Points, Diff, Place (empty for manual entry)
            col_letter = get_column_letter(col)
            
            # Points
            cell = ws[f'{col_letter}{row}']
            cell.value = ''
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = BORDER_THIN
            cell.font = FIGHTER_FONT
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            col += 1
            
            # Diff
            col_letter = get_column_letter(col)
            cell = ws[f'{col_letter}{row}']
            cell.value = ''
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = BORDER_THIN
            cell.font = FIGHTER_FONT
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            col += 1
            
            # Place
            col_letter = get_column_letter(col)
            cell = ws[f'{col_letter}{row}']
            cell.value = ''
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = BORDER_THIN
            cell.font = FIGHTER_FONT
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        return current_row + num_fighters + 1
    
    def _draw_finale_bracket(self, ws, pool_finalists, start_row):
        """
        Draw a horizontal cascade finale bracket with semifinals cascading to finals.
        
        Args:
            ws: Openpyxl worksheet
            pool_finalists: List of dicts with pool_name, rank_1, rank_2
            start_row: Starting row for finale bracket
        """
        match_height = 4  # 4 rows per match
        semi_col = 2  # Start at column B (shifted right)
        final_col = semi_col + 3  # Finals at column E
        
        # Headers on same row
        header_row = start_row
        ws.cell(row=header_row, column=semi_col, value='Semi-Finals')
        ws.cell(row=header_row, column=semi_col).font = HEADER_FONT
        ws.cell(row=header_row, column=semi_col).fill = HEADER_FILL
        ws.merge_cells(f'{get_column_letter(semi_col)}{header_row}:{get_column_letter(semi_col+1)}{header_row}')
        
        ws.cell(row=header_row, column=final_col, value='FINAL')
        ws.cell(row=header_row, column=final_col).font = HEADER_FONT
        ws.cell(row=header_row, column=final_col).fill = HEADER_FILL
        ws.merge_cells(f'{get_column_letter(final_col)}{header_row}:{get_column_letter(final_col+1)}{header_row}')
        
        pool1_name = pool_finalists[0].get('pool_name', 'Pool 1')
        pool2_name = pool_finalists[1].get('pool_name', 'Pool 2') if len(pool_finalists) > 1 else 'Pool 2'
        
        # Semi-finals - vertical stack (3 rows per match now: fighter, vs, fighter)
        semi1_row = header_row + 1
        semi2_row = semi1_row + 3 + 1  # 3 rows per match + 1 gap
        
        # SEMI 1: Pool 1 Winner vs Pool 2 Winner
        row = semi1_row
        cell = ws.cell(row=row, column=semi_col, value=f'{pool1_name} Winner')
        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        cell.border = BORDER_THIN
        cell.alignment = Alignment(horizontal='center')
        
        row += 1
        cell = ws.cell(row=row, column=semi_col, value='vs')
        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        cell.border = BORDER_THIN
        cell.font = Font(bold=True, color="FF0000")
        cell.alignment = Alignment(horizontal='center')
        
        row += 1
        cell = ws.cell(row=row, column=semi_col, value=f'{pool2_name} Winner')
        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        cell.border = BORDER_THIN
        cell.alignment = Alignment(horizontal='center')
        
        # SEMI 2: Pool 1 2nd vs Pool 2 2nd
        row = semi2_row
        cell = ws.cell(row=row, column=semi_col, value=f'{pool1_name} 2nd')
        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        cell.border = BORDER_THIN
        cell.alignment = Alignment(horizontal='center')
        
        row += 1
        cell = ws.cell(row=row, column=semi_col, value='vs')
        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        cell.border = BORDER_THIN
        cell.font = Font(bold=True, color="FF0000")
        cell.alignment = Alignment(horizontal='center')
        
        row += 1
        cell = ws.cell(row=row, column=semi_col, value=f'{pool2_name} 2nd')
        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        cell.border = BORDER_THIN
        cell.alignment = Alignment(horizontal='center')
        
        # FINAL: Center between Semi 1 and Semi 2, one line down
        final_row = semi1_row + 2  # One line lower than semi1_row + 1
        
        # Final - Fighter 1 (Semi 1 Winner)
        row = final_row
        cell = ws.cell(row=row, column=final_col, value='Semi 1 Winner')
        cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        cell.border = BORDER_THIN
        cell.alignment = Alignment(horizontal='center')
        
        row += 1
        cell = ws.cell(row=row, column=final_col, value='vs')
        cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        cell.border = BORDER_THIN
        cell.font = Font(bold=True, color="FF0000")
        cell.alignment = Alignment(horizontal='center')
        
        row += 1
        cell = ws.cell(row=row, column=final_col, value='Semi 2 Winner')
        cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        cell.border = BORDER_THIN
        cell.alignment = Alignment(horizontal='center')
