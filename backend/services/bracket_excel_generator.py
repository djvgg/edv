"""
Flexible bracket generator using Excel output + optional PDF export.
Uses round computation from v2 with snake seeding for flexible tournaments.
"""

import os
import logging
from typing import List, Dict, Tuple, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utils.logging import get_logger

# Import frontend styles
import sys
sys.path.insert(0, r'c:\UNI\TOP\edv_backend')
from frontend.styles import COLORS

logger = get_logger('bracket_excel_generator')

# Excel styling - using frontend COLORS
HEADER_FILL = PatternFill(start_color=COLORS['accent_blue'].lstrip('#'), end_color=COLORS['accent_blue'].lstrip('#'), fill_type="solid")
HEADER_FONT = Font(bold=True, color=COLORS['text_primary'].lstrip('#'), size=11)
MATCH_FILL = PatternFill(start_color=COLORS['bg_darker'].lstrip('#'), end_color=COLORS['bg_darker'].lstrip('#'), fill_type="solid")
MATCH_FONT = Font(size=10, color=COLORS['text_primary'].lstrip('#'))
VS_FONT = Font(bold=True, size=9, color=COLORS['accent_red'].lstrip('#'))
BORDER_THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
# Thick border for match boxes
BORDER_MATCH = Border(
    left=Side(style='medium'),
    right=Side(style='medium'),
    top=Side(style='medium'),
    bottom=Side(style='medium')
)
# Thick right border for connector
BORDER_CONNECTOR_RIGHT = Border(
    left=Side(style='thin'),
    right=Side(style='thick'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def _apply_snake_seeding(fighters):
    """Apply snake seeding to fighter list and pair them into matches."""
    if not fighters:
        return []
    
    seeded = []
    for i, fighter in enumerate(fighters):
        if isinstance(fighter, dict):
            name = fighter.get('Name', f'Fighter {i+1}')
            club = fighter.get('Verein', '')
        elif isinstance(fighter, (list, tuple)):
            name = fighter[0] if len(fighter) > 0 else f'Fighter {i+1}'
            club = fighter[1] if len(fighter) > 1 else ''
        else:
            name = str(fighter)
            club = ''
        seeded.append((name, club))
    
    # Pair fighters into matches for first round
    matches = []
    for i in range(0, len(seeded), 2):
        if i + 1 < len(seeded):
            p1, c1 = seeded[i]
            p2, c2 = seeded[i + 1]
            matches.append((p1, p2, c1, c2))  # 4-tuple: (name1, name2, club1, club2)
        else:
            # Odd fighter gets bye
            p1, c1 = seeded[i]
            matches.append((p1, 'BYE', c1, ''))
    
    return matches


def _compute_rounds(first_round_matches):
    """Compute all bracket rounds from first round matches."""
    if not first_round_matches:
        return []
    
    rounds = []
    current_round = list(first_round_matches)  # List of (p1, p2, c1, c2) tuples
    
    while len(current_round) > 1:
        rounds.append(current_round)
        next_round = []
        for i in range(0, len(current_round), 2):
            if i + 1 < len(current_round):
                # Two matches winner pairing: winner of match i vs winner of match i+1
                next_round.append(('TBD', 'TBD', 'TBD', 'TBD'))
            else:
                # Odd number, bye through - but keep 4-tuple structure
                match_data = current_round[i]
                p1 = match_data[0] if len(match_data) > 0 else 'TBD'
                c1 = match_data[2] if len(match_data) > 2 else ''
                next_round.append((p1, 'BYE', c1, ''))
        current_round = next_round
    
    # Final round
    if current_round:
        rounds.append(current_round)
    
    return rounds


def _compute_loser_rounds(wb_rounds):
    """
    Compute loser bracket rounds from winners bracket rounds.
    Uses same algorithm as viewer's _compute_loser_rounds_for_preview.
    
    Args:
        wb_rounds: List of winners bracket rounds (each match is 4-tuple)
    
    Returns:
        List of loser bracket rounds
    """
    if len(wb_rounds) < 2:
        return []
    
    def get_loser_from_tuple(match_tuple):
        """Extract loser from a 4-tuple (p1, p2, c1, c2). Since brackets are empty, return 'TBD'."""
        # In empty brackets, we don't know winners yet, so always return 'TBD'
        return 'TBD'
    
    loser_rounds = []
    
    # LB Round 0: pair consecutive losers from WB Round 0
    wb_r0_losers = [get_loser_from_tuple(m) for m in wb_rounds[0]]
    lb_r0_matches = []
    for i in range(0, len(wb_r0_losers), 2):
        p1 = wb_r0_losers[i]
        p2 = wb_r0_losers[i + 1] if i + 1 < len(wb_r0_losers) else 'Freilos'
        lb_r0_matches.append((p1, p2, '', ''))  # Empty clubs for losers
    loser_rounds.append(lb_r0_matches)
    
    # Subsequent LB rounds: mix LB winners with WB losers
    for r in range(1, len(wb_rounds)):
        wb_r_losers = [get_loser_from_tuple(m) for m in wb_rounds[r]]
        lb_r_prev_winners = ['TBD'] * len(loser_rounds[r - 1])  # All TBD for empty brackets
        
        if r == len(wb_rounds) - 1:
            # Final: LB winner vs WB final loser = "3rd Place"
            p1 = lb_r_prev_winners[0] if lb_r_prev_winners else 'TBD'
            p2 = wb_r_losers[0] if wb_r_losers else 'TBD'
            loser_rounds.append([(p1, p2, '', '')])
        else:
            # Injection or reduction round
            lb_matches = []
            prev_count = len(loser_rounds[r - 1])
            curr_wb_losers = len(wb_r_losers)
            
            if curr_wb_losers >= prev_count:
                # Injection: 1-to-1 matching
                for i in range(prev_count):
                    p1 = lb_r_prev_winners[i] if i < len(lb_r_prev_winners) else 'TBD'
                    p2 = wb_r_losers[i] if i < len(wb_r_losers) else 'Freilos'
                    lb_matches.append((p1, p2, '', ''))
            else:
                # Reduction: pair up LB winners, they fight each other
                for i in range(0, len(lb_r_prev_winners), 2):
                    p1 = lb_r_prev_winners[i] if i < len(lb_r_prev_winners) else 'TBD'
                    p2 = lb_r_prev_winners[i + 1] if i + 1 < len(lb_r_prev_winners) else 'Freilos'
                    lb_matches.append((p1, p2, '', ''))
            
            loser_rounds.append(lb_matches)
    
    return loser_rounds


def _compute_finale_bracket(wb_rounds, lb_rounds):
    """
    Compute finale bracket from winners and loser bracket finalists.
    Creates a small 4-fighter bracket with semi-finals and finals.
    
    Args:
        wb_rounds: Winners bracket rounds (get finalist from last round)
        lb_rounds: Loser bracket rounds (get finalist from last round)
    
    Returns:
        List of finale rounds (2 rounds: semi-finals, finals)
    """
    if not wb_rounds or not lb_rounds:
        return []
    
    # Extract finalists
    # Winners bracket finalist (first fighter from final match)
    wb_finalist = wb_rounds[-1][0][0] if wb_rounds[-1] else 'TBD'
    wb_finalist_club = wb_rounds[-1][0][2] if wb_rounds[-1] else ''
    
    # Loser bracket finalist (first fighter from final match)
    lb_finalist = lb_rounds[-1][0][0] if lb_rounds[-1] else 'TBD'
    lb_finalist_club = lb_rounds[-1][0][2] if lb_rounds[-1] else ''
    
    # Secondary finalists (for 4-fighter bracket) - take losers from semi-finals in WB
    semi_finalists_1 = wb_rounds[-2][0] if len(wb_rounds) > 1 else ('TBD', 'TBD', '', '')
    semi_finalists_2 = wb_rounds[-2][1] if len(wb_rounds) > 1 and len(wb_rounds[-2]) > 1 else ('TBD', 'TBD', '', '')
    
    # Semi-finals: 2 matches
    # Match 1: WB finalist vs first semi-finalist
    # Match 2: LB finalist vs second semi-finalist
    semi_finals = [
        (semi_finalists_1[0], semi_finalists_2[0], semi_finalists_1[2], semi_finalists_2[2]),
        (wb_finalist, lb_finalist, wb_finalist_club, lb_finalist_club)
    ]
    
    # Finals: winners of semi-finals (TBD)
    finals = [
        ('TBD', 'TBD', '', '')
    ]
    
    return [semi_finals, finals]


class BracketExcelGenerator:
    """Generate tournament brackets in Excel format with flexible layout."""
    
    def __init__(self):
        self.logger = logger
    
    def generate_bracket_excel(self, 
                              output_path: str,
                              fighters: List[Dict],
                              bracket_type: str = 'double',
                              title: str = "Tournament Bracket") -> bool:
        """
        Generate Excel bracket file with horizontal cascade layout.
        
        Args:
            output_path: Path to save Excel file
            fighters: List of fighter dicts with 'Name' and 'Verein' (club)
            bracket_type: 'single' or 'double' elimination
            title: Tournament title
        
        Returns:
            True if successful
        """
        try:
            os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
            
            # Apply snake seeding
            bracket_pairs = _apply_snake_seeding(fighters)
            
            # Compute rounds
            rounds = _compute_rounds(bracket_pairs)
            if not rounds:
                self.logger.error("Failed to compute bracket rounds")
                return False
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Bracket"
            
            # Set up sheet properties
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.orientation = 'landscape'
            
            # Title
            ws['A1'] = title
            ws['A1'].font = Font(bold=True, size=14)
            
            # Column setup per round (each round gets X columns)
            for col in range(1, 100):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            # Draw horizontal cascade bracket
            max_row_wb = self._draw_horizontal_bracket(ws, rounds, 3)
            
            # Draw loser bracket if double elimination
            if bracket_type == 'double':
                loser_rounds = _compute_loser_rounds(rounds)
                if loser_rounds:
                    # Start loser bracket below winners bracket (no gap)
                    loser_start_row = max_row_wb + 2
                    max_row_lb = self._draw_horizontal_bracket(ws, loser_rounds, loser_start_row, is_loser_bracket=True)
                    
                    # Draw finale bracket (top 4 finalists)
                    finale_rounds = _compute_finale_bracket(rounds, loser_rounds)
                    if finale_rounds:
                        finale_start_row = max_row_lb + 3
                        self._draw_horizontal_bracket(ws, finale_rounds, finale_start_row, is_finale_bracket=True)
            
            # Save
            wb.save(output_path)
            self.logger.info(f"✓ Bracket Excel file created: {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error creating bracket Excel: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _draw_horizontal_bracket(self, ws, rounds, start_row, is_loser_bracket=False, is_finale_bracket=False):
        """Draw bracket in true tree structure with proper cascade.
        
        Args:
            ws: Openpyxl worksheet
            rounds: List of rounds (each round is list of 4-tuples)
            start_row: Starting row for bracket
            is_loser_bracket: If True, use loser bracket positioning and labels
            is_finale_bracket: If True, use finale bracket labels (semi-finals/final)
            
        Returns:
            Max row used (for positioning next bracket)
        """
        
        # Calculate row positions for each match
        match_height = 4
        col = 1
        round_col_map = {}  # Map round -> starting column
        
        # For each round, calculate starting column
        for round_num in range(len(rounds)):
            round_col_map[round_num] = col
            col += 3  # 2 columns for match+club, 1 for connector
        
        # Row planning - different logic for loser bracket
        row_map = {}  # Map (round, match_num) -> starting_row
        y_gap = 1  # Extra spacing between matches in loser bracket
        
        if is_loser_bracket:
            # Loser bracket positioning: vertical stack in R0, then cascade
            current_row = start_row + 1
            for match_num in range(len(rounds[0])):
                row_map[(0, match_num)] = current_row
                current_row += match_height + y_gap
            
            # Subsequent rounds: reduction or injection
            for round_num in range(1, len(rounds)):
                prev_count = len(rounds[round_num - 1])
                curr_count = len(rounds[round_num])
                
                for match_num in range(curr_count):
                    if curr_count < prev_count:
                        # Reduction: center between two predecessors
                        ya = row_map.get((round_num - 1, match_num * 2), start_row + 1) + match_height // 2
                        yb = row_map.get((round_num - 1, match_num * 2 + 1), ya) + match_height // 2
                        y = ((ya + yb) // 2) - match_height // 2
                    else:
                        # Injection: same row as predecessor
                        ya = row_map.get((round_num - 1, match_num), start_row + 1) + match_height // 2
                        y = ya - match_height // 2
                    row_map[(round_num, match_num)] = y
        else:
            # Winners bracket positioning: cascade from first round
            current_row = start_row + 1
            for match_num in range(len(rounds[0])):
                row_map[(0, match_num)] = current_row
                current_row += match_height
            
            # For subsequent rounds, place at vertical center of predecessors
            for round_num in range(1, len(rounds)):
                for match_num in range(len(rounds[round_num])):
                    prev_match1_row = row_map[(round_num - 1, match_num * 2)]
                    if match_num * 2 + 1 < len(rounds[round_num - 1]):
                        prev_match2_row = row_map[(round_num - 1, match_num * 2 + 1)]
                    else:
                        prev_match2_row = prev_match1_row
                    
                    center_row = (prev_match1_row + prev_match2_row) // 2
                    row_map[(round_num, match_num)] = center_row
        
        # Draw connector lines between rounds
        for round_num in range(len(rounds) - 1):
            from_col = round_col_map[round_num] + 2  # Connector column
            
            for match_num in range(len(rounds[round_num])):
                if (round_num, match_num) not in row_map:
                    continue
                    
                from_row = row_map[(round_num, match_num)] + 1  # Middle of match (vs row)
                
                # Determine target match in next round
                if is_loser_bracket and len(rounds[round_num]) > len(rounds[round_num + 1]):
                    # Reduction: two source matches go to one target
                    to_match_num = match_num // 2
                else:
                    # Normal injection
                    to_match_num = match_num
                
                if (round_num + 1, to_match_num) in row_map:
                    to_row = row_map[(round_num + 1, to_match_num)] + 1  # Middle of next match
                    
                    # Add thick borders on connector column
                    connector_col = from_col
                    for r in range(min(from_row, to_row), max(from_row, to_row) + 1):
                        cell = ws.cell(row=r, column=connector_col)
                        if cell.border:
                            existing = cell.border
                            cell.border = Border(
                                left=existing.left,
                                right=Side(style='thick'),
                                top=existing.top,
                                bottom=existing.bottom
                            )
                        else:
                            cell.border = BORDER_CONNECTOR_RIGHT
        
        # Now draw all matches
        max_row = start_row
        for round_num, matches in enumerate(rounds):
            col = round_col_map[round_num]
            
            # Round header
            if is_finale_bracket:
                if round_num == 0:
                    round_label = "Semi-Finals"
                else:
                    round_label = "FINAL"
            elif is_loser_bracket:
                if round_num == 0:
                    round_label = "Loser R1"
                elif round_num == len(rounds) - 1:
                    round_label = "3rd Place"
                else:
                    round_label = f"Loser R{round_num + 1}"
            else:
                if round_num == 0:
                    round_label = "Round 1"
                elif round_num == len(rounds) - 1:
                    round_label = "FINAL"
                else:
                    round_label = f"Round {round_num + 1}"
            
            header_row = start_row
            header_cell = ws.cell(row=header_row, column=col, value=round_label)
            header_cell.font = HEADER_FONT
            header_cell.fill = HEADER_FILL
            ws.merge_cells(f'{get_column_letter(col)}{header_row}:{get_column_letter(col+1)}{header_row}')
            
            # Draw each match
            for match_num, match_data in enumerate(matches):
                if (round_num, match_num) not in row_map:
                    continue
                    
                row = row_map[(round_num, match_num)]
                
                # Parse match data - should be 4-tuple (name1, name2, club1, club2)
                if isinstance(match_data, (list, tuple)) and len(match_data) == 4:
                    p1, p2, club1, club2 = match_data
                else:
                    p1, p2, club1, club2 = 'ERROR', 'ERROR', 'BAD', 'DATA'
                
                # Row: Fighter 1 name
                cell = ws.cell(row=row, column=col, value=p1)
                cell.fill = MATCH_FILL
                cell.font = MATCH_FONT
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = BORDER_MATCH
                ws.row_dimensions[row].height = 16
                
                # Club 1
                cell = ws.cell(row=row, column=col + 1, value=club1)
                cell.fill = MATCH_FILL
                cell.font = Font(size=8, italic=True)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = BORDER_MATCH
                
                # Connector column (vertical line)
                cell = ws.cell(row=row, column=col + 2)
                cell.border = BORDER_CONNECTOR_RIGHT
                
                row += 1
                
                # Row: "vs"
                cell = ws.cell(row=row, column=col, value="vs")
                cell.font = Font(bold=True, size=9, color="FF0000")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = BORDER_MATCH
                ws.row_dimensions[row].height = 14
                
                cell = ws.cell(row=row, column=col + 1)
                cell.border = BORDER_MATCH
                
                # Connector
                cell = ws.cell(row=row, column=col + 2)
                cell.border = BORDER_CONNECTOR_RIGHT
                
                row += 1
                
                # Row: Fighter 2 name
                cell = ws.cell(row=row, column=col, value=p2)
                cell.fill = MATCH_FILL
                cell.font = MATCH_FONT
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = BORDER_MATCH
                ws.row_dimensions[row].height = 16
                
                # Club 2
                cell = ws.cell(row=row, column=col + 1, value=club2)
                cell.fill = MATCH_FILL
                cell.font = Font(size=8, italic=True)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = BORDER_MATCH
                
                # Connector
                cell = ws.cell(row=row, column=col + 2)
                cell.border = BORDER_CONNECTOR_RIGHT
                
                # Track max row
                max_row = max(max_row, row)
        
        return max_row
    
    def generate_bracket_with_pdf(self,
                                  output_excel_path: str,
                                  output_pdf_path: str,
                                  fighters: List[Dict],
                                  bracket_type: str = 'double',
                                  title: str = "Tournament Bracket") -> Tuple[bool, bool]:
        """
        Generate both Excel and PDF versions.
        
        Returns:
            Tuple (excel_success, pdf_success)
        """
        excel_ok = self.generate_bracket_excel(output_excel_path, fighters, bracket_type, title)
        
        # Try to export to PDF via Excel COM if available
        pdf_ok = False
        try:
            import win32com.client
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            
            wb = excel.Workbooks.Open(os.path.abspath(output_excel_path))
            wb.ExportAsFixedFormat(0, os.path.abspath(output_pdf_path))  # 0 = PDF format
            wb.Close()
            excel.Quit()
            
            self.logger.info(f"✓ Bracket PDF file created: {output_pdf_path}")
            pdf_ok = True
        except ImportError:
            self.logger.warning("pywin32 not available, skipping PDF export")
        except Exception as e:
            self.logger.warning(f"PDF export failed: {e}")
        
        return excel_ok, pdf_ok


# Convenience functions
def create_bracket_excel(output_path: str,
                         fighters: List[Dict],
                         bracket_type: str = 'double',
                         title: str = "Tournament Bracket") -> bool:
    """Create Excel bracket file."""
    gen = BracketExcelGenerator()
    return gen.generate_bracket_excel(output_path, fighters, bracket_type, title)


def create_bracket_excel_and_pdf(excel_path: str,
                                 pdf_path: str,
                                 fighters: List[Dict],
                                 bracket_type: str = 'double',
                                 title: str = "Tournament Bracket") -> Tuple[bool, bool]:
    """Create both Excel and PDF bracket files."""
    gen = BracketExcelGenerator()
    return gen.generate_bracket_with_pdf(excel_path, pdf_path, fighters, bracket_type, title)
